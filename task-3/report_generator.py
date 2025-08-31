"""Excel report generation for Cricket Fielding Analysis.

Creates a multi-sheet workbook:
- Ball-by-Ball: raw/clean dataset
- Performance Matrix: per-player counts + PS
- Weights & Formula: weights table and PS formula reference
"""

from __future__ import annotations

from typing import Dict

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


def _autosize(ws) -> None:
	for col in ws.columns:
		max_length = 0
		col_letter = get_column_letter(col[0].column)
		for cell in col:
			try:
				value_len = len(str(cell.value)) if cell.value is not None else 0
				max_length = max(max_length, value_len)
			except Exception:
				pass
		ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def _build_pretty_matrix(summary_df: pd.DataFrame) -> pd.DataFrame:
	"""Return matrix with columns ordered and labeled like the sample."""
	ordered = summary_df.copy()
	# Ensure the canonical order and friendly headers
	ordered = ordered[["CP", "GT", "C", "DC", "ST", "RO", "MRO", "DH", "RS", "PS"]]
	ordered = ordered.rename(
		columns={
			"CP": "Clean Picks (CP)",
			"GT": "Good Throws (GT)",
			"C": "Catches (C)",
			"DC": "Dropped Catches (DC)",
			"ST": "Stumpings (ST)",
			"RO": "Run Outs (RO)",
			"MRO": "Missed Run Outs (MR)",
			"DH": "Direct Hits (DH)",
			"RS": "Runs Saved (RS)",
			"PS": "Performance Score (PS)",
		}
	)
	return ordered


def _write_calculations_sheet(writer, summary_df: pd.DataFrame, weights_obj) -> None:
	"""Add a sheet showing PS formula expansion per player."""
	ws_name = "Calculations"
	calc_df = pd.DataFrame({"Notes": [
		"PS=(CP×WCP)+(GT×WGT)+(C×WC)+(DC×WDC)+(ST×WST)+(RO×WRO)+(MRO×WMRO)+(DH×WDH)+RS",
		"Where W are weights and RS is added directly.",
	]})
	calc_df.to_excel(writer, index=False, header=False, sheet_name=ws_name, startrow=0)
	ws = writer.book[ws_name]
	row = 4
	for player, rowvals in summary_df.iterrows():
		cp, gt, c, dc, st, ro, mro, dh, rs, ps = (
			rowvals["CP"], rowvals["GT"], rowvals["C"], rowvals["DC"], rowvals["ST"],
			rowvals["RO"], rowvals["MRO"], rowvals["DH"], rowvals["RS"], rowvals["PS"],
		)
		expr = (
			f"PS=({cp}×{weights_obj.WCP})+({gt}×{weights_obj.WGT})+({c}×{weights_obj.WC})+"
			f"({dc}×{weights_obj.WDC})+({st}×{weights_obj.WST})+({ro}×{weights_obj.WRO})+"
			f"({mro}×{weights_obj.WMRO})+({dh}×{weights_obj.WDH})+{rs}"
		)
		ws.cell(row=row, column=1, value=player)
		ws.cell(row=row + 1, column=1, value=expr)
		ws.cell(row=row + 2, column=1, value=f"PS={ps}")
		row += 4
	_autosize(ws)


def create_excel_report(ball_by_ball: pd.DataFrame, summary_df: pd.DataFrame, weights_obj, output_path: str) -> None:
	"""Write a formatted Excel report to the given path."""
	with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

		# Sheet 1: Ball-by-Ball with legend header band (like sample)
		ball_by_ball.to_excel(writer, index=False, sheet_name="Ball-by-Ball", startrow=5)
		ws1 = writer.book["Ball-by-Ball"]
		# Legend rows
		head_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		bold = Font(bold=True)
		legend_rows = [
			["Pick", "Y->", "Clean Pick", "N->", "Fumble", "C->", "Catch", "DC->", "Dropped Catch"],
			["Throw", "Y->", "Good Throw", "N->", "Bad throw", "RO->", "Run Out", "MR->", "Missed Runout", "ST->", "Stumping"],
			["Runs", "Use '+' for runs saved, '-' for conceded"],
		]
		for r_idx, row_vals in enumerate(legend_rows, start=1):
			for c_idx, val in enumerate(row_vals, start=1):
				cell = ws1.cell(row=r_idx, column=c_idx, value=val)
				cell.font = bold
				cell.alignment = Alignment(horizontal="left")
				cell.fill = head_fill
		# Freeze panes below legend and column headers
		ws1.freeze_panes = "A7"
		_autosize(ws1)

		# Sheet 2: Performance Matrix (counts + PS)
		matrix = summary_df.copy()
		matrix.index.name = "Player Name"
		pretty = _build_pretty_matrix(matrix)
		pretty.reset_index().to_excel(writer, index=False, sheet_name="Performance Matrix")
		ws2 = writer.book["Performance Matrix"]
		_autosize(ws2)
		ws2.freeze_panes = "A2"

		# Sheet 3: Weights & Formula
		weights_dict: Dict[str, float] = {
			"WCP": weights_obj.WCP,
			"WGT": weights_obj.WGT,
			"WC": weights_obj.WC,
			"WDC": weights_obj.WDC,
			"WST": weights_obj.WST,
			"WRO": weights_obj.WRO,
			"WMRO": weights_obj.WMRO,
			"WDH": weights_obj.WDH,
		}
		weights_df = pd.DataFrame({"Weight": list(weights_dict.keys()), "Value": list(weights_dict.values())})
		weights_df.to_excel(writer, index=False, sheet_name="Weights & Formula", startrow=0)
		ws3 = writer.book["Weights & Formula"]
		# Add formula text
		start_row = len(weights_df) + 3
		ws3.cell(row=start_row, column=1, value="PS Formula:")
		ws3.cell(
			row=start_row + 1,
			column=1,
			value="PS = (CP×WCP) + (GT×WGT) + (C×WC) + (DC×WDC) + (ST×WST) + (RO×WRO) + (MRO×WMRO) + (DH×WDH) + RS",
		)
		ws3["A1"].font = Font(bold=True)
		ws3["A1"].alignment = Alignment(horizontal="left")
		_autosize(ws3)

		# Sheet 4: Calculations (per-player expansion)
		_write_calculations_sheet(writer, matrix, weights_obj)


__all__ = ["create_excel_report"]


